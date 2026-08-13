#!/usr/bin/env python3
"""cosmotherm_postprocess.py — rank a single COSMOtherm solubility screen.

Reads molecules/<mol>/<protocol>/cosmotherm/screen.tab — one "Solubility Solvent
Screening" job containing BOTH pure solvents and binary mixtures, each with
log10(x_RS) on one relative (DG_fus-free) scale — and writes a ranking:

  <mol>-screening-ranking.xlsx   ranked table (pure + mixture, one scale)
  results.csv                    same, long-form

Pure rows have a single-word Solvent; mixture rows carry a multi-word label like
'h2o propanol X={ 0.1790 0.8210 }'. Both sit on the same log10(x_RS) axis, exactly
as in the COSMOthermX desktop screen (no water reference, no DG_fus).

Parsing: COSMOtherm mixes a LEFT-aligned Solvent column with RIGHT-aligned numeric
columns, and wide numbers overflow the header word — so fixed-width slicing fails.
Instead we split each data row on runs of >=2 spaces: 'Nr Solvent' (single-space
internals, even for mixtures) stays one chunk; the numeric columns separate cleanly.

Usage: cosmotherm_postprocess.py <molecule> <dft-protocol>
"""
from __future__ import annotations
import csv
import re
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

HEADER_RE = re.compile(r'^\s*Nr\s+Solvent\b')
META_RE = re.compile(r'^\s*(Property|Settings|Units|General)\b')
GAP_RE = re.compile(r'\s{2,}')


def _to_float(v):
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def parse_screen_tab(path: Path) -> list[dict]:
    """Return [{Nr, Solvent, log10(x_RS), w_RS, log10(S_RS)}, ...]."""
    lines = [l.rstrip('\n') for l in path.open()]
    hdr_i = next((i for i, l in enumerate(lines) if HEADER_RE.match(l)), None)
    if hdr_i is None:
        return []
    recs = []
    for l in lines[hdr_i + 1:]:
        s = l.strip()
        if not s or META_RE.match(l) or HEADER_RE.match(l):
            continue
        chunks = GAP_RE.split(s)                 # split on runs of >=2 spaces
        head = chunks[0].split(' ', 1)           # "Nr Solvent" (single space)
        if not head[0].isdigit():
            continue
        recs.append({
            'Nr': head[0],
            'Solvent': head[1].strip() if len(head) > 1 else '',
            'log10(x_RS)': chunks[1] if len(chunks) > 1 else '',
            'w_RS': chunks[2] if len(chunks) > 2 else '',
            'log10(S_RS)': chunks[3] if len(chunks) > 3 else '',
        })
    return recs


def main() -> int:
    if len(sys.argv) < 3:
        print(f"Usage: {Path(sys.argv[0]).name} <molecule> <dft-protocol>", file=sys.stderr)
        return 2
    molecule, protocol = sys.argv[1], sys.argv[2]
    work = Path(__file__).resolve().parent.parent / 'molecules' / molecule / protocol / 'cosmotherm'
    tab = work / 'screen.tab'
    if not tab.is_file():
        print(f"ERROR: {tab} not found — run cosmotherm_screen.sh first", file=sys.stderr)
        return 2

    rows = parse_screen_tab(tab)
    if not rows:
        print(f"ERROR: no data rows parsed from {tab}", file=sys.stderr)
        return 1

    recs = []
    for r in rows:
        solv = r['Solvent']
        recs.append({
            'system': solv,
            'type': 'mixture' if 'x={' in solv.lower() else 'pure',
            'log10(x_RS)': _to_float(r['log10(x_RS)']),
            'w_RS': r['w_RS'],
            'log10(S_RS)': _to_float(r['log10(S_RS)']),
        })
    recs.sort(key=lambda x: x['log10(x_RS)'] if x['log10(x_RS)'] is not None else float('-inf'),
              reverse=True)

    n_pure = sum(1 for r in recs if r['type'] == 'pure')
    n_mix = len(recs) - n_pure
    print(f"[screen] {len(recs)} systems parsed ({n_pure} pure, {n_mix} mixtures)")

    # --- CSV ---
    csv_path = work / 'results.csv'
    with csv_path.open('w', newline='') as f:
        w = csv.writer(f)
        w.writerow(['rank', 'system', 'type', 'log10(x_RS)', 'w_RS', 'log10(S_RS)'])
        for i, r in enumerate(recs, 1):
            w.writerow([i, r['system'], r['type'], r['log10(x_RS)'], r['w_RS'], r['log10(S_RS)']])
    print(f"wrote {csv_path}")

    # --- XLSX ---
    if HAS_OPENPYXL:
        wb = Workbook(); ws = wb.active; ws.title = "Screening ranking"
        bold = Font(bold=True); grey = PatternFill("solid", fgColor="EEEEEE")
        note = ("Relative solubility screen — log10(x_RS), DG_fus-free. Pure solvents + binary "
                "mixtures on ONE scale, sorted most->least soluble. Mixture labels show mole "
                "fractions (v/v converted). Ranks COSMO-RS affinity, not absolute solubility; "
                "magnitudes comparable within this solute only.")
        ws.cell(row=1, column=1, value=note).font = Font(italic=True)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        headers = ['rank', 'system', 'type', 'log10(x_RS)', 'w_RS', 'log10(S_RS)']
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=3, column=ci, value=h); c.font = bold; c.fill = grey
        for i, r in enumerate(recs, 1):
            ws.cell(row=3 + i, column=1, value=i)
            ws.cell(row=3 + i, column=2, value=r['system'])
            ws.cell(row=3 + i, column=3, value=r['type'])
            ws.cell(row=3 + i, column=4, value=r['log10(x_RS)'])
            ws.cell(row=3 + i, column=5, value=r['w_RS'])
            ws.cell(row=3 + i, column=6, value=r['log10(S_RS)'])
        ws.column_dimensions['B'].width = 36
        for col in ('A', 'C', 'D', 'E', 'F'):
            ws.column_dimensions[col].width = 15
        out = work / f"{molecule}-screening-ranking.xlsx"
        wb.save(out)
        print(f"wrote {out}")
    else:
        print("NOTE: openpyxl not installed — xlsx skipped (csv written).", file=sys.stderr)
    return 0


if __name__ == '__main__':
    sys.exit(main())
